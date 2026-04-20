import openpyxl

def extraer_estudiantes_de_sede(workbook, sheet_name, n_estudiantes, inicio_fila_relativa):
    """
    Lee los datos de los estudiantes de una hoja específica de Excel.
    
    :param workbook: Objeto de openpyxl.
    :param sheet_name: Nombre de la hoja (sede).
    :param n_estudiantes: Cantidad de alumnos a procesar.
    :param inicio_fila_relativa: El índice relativo de inicio (empieza en 1 según la GUI).
    :return: Lista de diccionarios con datos de estudiantes.
    """
    estudiantes = []
    sheet = workbook[sheet_name]
    
    # La lógica original dice que los datos empiezan en la fila 11.
    # El limite_filas se calculaba como n_estudiantes + 11.
    # El bucle original era: range(inicio_estudiante, limite_filas - 1)
    # donde inicio_estudiante = item["inicio"] + 1 (que viene de la fila 11).
    
    # Ajuste de índices para coincidir con la lógica original
    start_row = 11 + (inicio_fila_relativa - 1)
    end_row = start_row + n_estudiantes
    
    # Leer columnas necesarias en el rango de filas
    # B a M (2 a 13): Matrícula
    # N a P (14 a 16): Nombres
    # Q a T (17 a 20): Materias
    
    columnas_matricula = list(sheet.iter_cols(min_col=3, max_col=14, min_row=11, max_row=end_row))
    columnas_nombres = list(sheet.iter_cols(min_col=15, max_col=17, min_row=11, max_row=end_row))
    columnas_materias = list(sheet.iter_cols(min_col=18, max_col=21, min_row=11, max_row=end_row))

    # El offset es porque iter_cols con min_row=11 devuelve celdas donde el índice 0 corresponde a la fila 11
    for i in range(inicio_fila_relativa - 1, inicio_fila_relativa - 1 + n_estudiantes):
        try:
            # Extraer Matrícula
            matricula_parts = [str(columnas_matricula[c][i].value) for c in range(len(columnas_matricula))]
            matricula_str = "".join(matricula_parts).replace("None", "").strip()
            
            if not matricula_str:
                continue

            # Extraer Nombre
            nombres_parts = [str(columnas_nombres[c][i].value) for c in range(len(columnas_nombres))]
            nombre_completo = " ".join(nombres_parts).replace("None", "").strip()

            # Extraer Materias
            materias = [str(columnas_materias[c][i].value) for c in range(len(columnas_materias)) 
                       if columnas_materias[c][i].value and str(columnas_materias[c][i].value) != 'None']

            estudiantes.append({
                "matricula": matricula_str,
                "nombre": nombre_completo,
                "materias": materias,
                "fila_excel": 11 + i,
                "sede": sheet_name
            })
        except IndexError:
            # Fin de los datos alcanzado antes de lo esperado
            break
            
    return estudiantes
