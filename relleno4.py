import openpyxl
import pyautogui
import time
import sys
from rich.console import Console
from rich.table import Table
from rich.panel import Panel
from rich.prompt import Prompt, IntPrompt
from rich.progress import track

console = Console()

def main():
    console.print(Panel.fit(
        "[bold blue]SIOSAD - Sistema de Captura Automática[/bold blue]\n"
        "[white]Asegúrese de tener el SIOSAD abierto en la pantalla de solicitud de exámenes.[/white]",
        title="Bienvenido", border_style="green"
    ))

    # Carga de Archivo con validación
    while True:
        excel_name = Prompt.ask("[bold yellow]Ingrese el nombre del archivo Excel (sin extensión)[/bold yellow]")
        excel_file = f"{excel_name}.xlsx"
        try:
            with console.status("[bold green]Cargando base de datos..."):
                workbook = openpyxl.load_workbook(excel_file, data_only=True)
            break
        except FileNotFoundError:
            console.print(f"[bold red]Error:[/bold red] El archivo '{excel_file}' no existe. Reintente.")

    # Iterar sobre hojas
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        
        console.print(f"\n[bold magenta]Sede Actual:[/bold magenta] {sheet_name}")

        # Validaciones de entrada de usuario
        try:
            n_estudiantes = IntPrompt.ask(f"¿Cuántos estudiantes procesar en la sede {sheet_name}?", default=0)
            if n_estudiantes <= 0:
                console.print("[yellow]Sede omitida.[/yellow]")
                continue

            inicio_input = Prompt.ask("Iniciar desde el estudiante # (Enter para empezar desde el 1)", default="1")
            inicio_estudiante = int(inicio_input) + 1
        except ValueError:
            console.print("[bold red]Entrada inválida. Saltando sede...[/bold red]")
            continue

        limite_filas = n_estudiantes + 11
        
        # Lectura de rangos
        columnas_a_leer = list(sheet.iter_cols(min_col=2, max_col=13, min_row=11, max_row=limite_filas))
        datos_estudiante = list(sheet.iter_cols(min_col=14, max_col=16, min_row=11, max_row=limite_filas))
        columnas_materias = list(sheet.iter_cols(min_col=17, max_col=20, min_row=11, max_row=limite_filas))

        # Procesamiento de filas
        for row_idx in range(inicio_estudiante, limite_filas - 1):
            try:
                # Extraer Datos
                matricula_lista = [str(columnas_a_leer[c][row_idx - 1].value) for c in range(len(columnas_a_leer))]
                matricula_str = "".join(matricula_lista).replace("None", "")
                
                nombres_lista = [str(datos_estudiante[c][row_idx - 1].value) for c in range(len(datos_estudiante))]
                nombre_completo = " ".join(nombres_lista).replace("None", "").strip()

                materias = [str(columnas_materias[c][row_idx - 1].value) for c in range(len(columnas_materias)) 
                           if columnas_materias[c][row_idx - 1].value and str(columnas_materias[c][row_idx - 1].value) != 'None']

                # Mostrar Tabla de datos actuales
                table = Table(title=f"Registro actual: {row_idx - 1}")
                table.add_column("Campo", style="cyan")
                table.add_column("Valor", style="white")
                table.add_row("Estudiante", nombre_completo)
                table.add_row("Matrícula", matricula_str)
                table.add_row("Materias", ", ".join(materias))
                console.print(table)

                # Automatización PyAutoGUI
                console.print("[yellow]Iniciando captura en SIOSAD en 2 segundos...[/yellow]")
                time.sleep(2)
                
                pyautogui.click(x=150, y=150)
                pyautogui.write(matricula_str)
                pyautogui.press('enter')
                pyautogui.press('enter')
                pyautogui.write('404') 
                pyautogui.press('enter')
                pyautogui.write('S')
                pyautogui.write(str(sheet_name))
                pyautogui.press('enter')
                pyautogui.press('enter')

                for materia in materias:
                    pyautogui.write(materia)
                    pyautogui.press('enter')

                # Confirmación manual
                Prompt.ask("\n[bold cyan]Revise los datos en SIOSAD. Presione ENTER para GUARDAR (F2)[/bold cyan]")
                time.sleep(1)                
                pyautogui.press('f2')
                pyautogui.press('enter')
                pyautogui.press('enter')
                
                console.print("[bold green]✔ CAPTURA EXITOSA[/bold green]")
                
                opcion = Prompt.ask("Presione ENTER para el siguiente o 'q' para salir", default="")
                if opcion.lower() == 'q':
                    sys.exit()
                time.sleep(1)                    
                pyautogui.press('f9')


            except Exception as e:
                console.print(f"[bold red]Error en fila {row_idx}:[/bold red] {e}")
                break

    workbook.close()
    console.print(Panel("[bold green]Proceso finalizado correctamente.[/bold green]"))

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        console.print("\n[bold red]Programa detenido por el usuario.[/bold red]")